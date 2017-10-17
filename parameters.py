from math import *
from openpyxl import load_workbook
from openpyxl import Workbook
import logging

logging.info('#############################')
logging.info('##### Parametros Globais ####')
logging.info('#############################')

#Velocidade do vento [m/s]
u_inf = 10
#TSR
tsr = 18
#Raio do aerogerador [m]
r = 0.64
#Velocidade de rotacao 
omega = tsr*u_inf/r 
#Densidade do ar [kg/m3]
ro = 1.19
#Viscodidade do ar [kg/ms] 
mu = 0.0000184
#Area
a = 1
#Corda [m]
c = 0.032
#Cd0
cd_0 = 1
#Cl0
cl_0 = 1
#Epsilon
epsilon = 0.01
#Potencia Maxima
p_max = 0.5*ro*a*(u_inf**3)
#Pressao Atmosferica [Pa]
p_atm = 101325

logging.info('u_inf = ' + str(u_inf))
logging.info('tsr = ' + str(tsr))
logging.info('r = ' + str(r))
logging.info('omega = ' + str(omega))
logging.info('ro = ' + str(ro))
logging.info('mu = ' + str(mu))
logging.info('c = ' + str(c))
logging.info('epsilon = ' + str(epsilon))
logging.info('p_max = ' + str(p_max))
logging.info('p_atm = ' + str(p_atm))

logging.info('####################')
logging.info('##### Inputs  ######')
logging.info('####################')

n = 360
dteta = 2*pi/n

logging.info('n = ' + str(180))
logging.info('dteta = ' + str(dteta))


logging.info('####################')
logging.info('### Perfil NACA ####')
logging.info('####################')


logging.info('Abrindo documento com dados NACA')
wb = load_workbook(filename = 'dadosNACA.xlsx')
ws = wb[wb.sheetnames[0]]
ws_range = str.split(ws.dimensions,':')

#Define o Numero de valores de alfa e o numero pares (cd,cl)
alfa_range = int(ws_range[1][1:])
naca_range = ((ord(ws_range[1][:1].lower()) - 96)-1)/2

listaRe = []
dictNACA = {}
for naca in range(1, naca_range + 1):
    num_naca = str(ws[chr(97+2*naca).upper()+str(1)].value)
    re = str(ws[chr(97+2*naca).upper()+str(2)].value)
    if num_naca not in dictNACA:
        dictNACA[num_naca]={}
        if re not in dictNACA[num_naca]:
            dictNACA[num_naca][re]={}
    else:
        if re not in dictNACA[num_naca]:
            dictNACA[num_naca][re]={}
    if not listaRe.__contains__(re):
        listaRe.append(int(re))

listaRe.sort

# dictNACA[num_naca][re]={'alfa':[], 'cd': [], 'cl':[]}
# dictNACA[num_naca][re][alfa]

#Estrutura excel
#Linha 1: Valores Naca
#Linha 2: Valores Reynolds
#Linha 3: Titulos
#Coluna 1: Valores Alfa
for naca in range(1,naca_range+1):
    num_naca = str(ws[chr(97+2*naca).upper()+str(1)].value)
    re = str(ws[chr(97+2*naca).upper()+str(2)].value)
    dictNACA[num_naca][re]={'alfa':[],'cd':[],'cl':[]}
    for i in range(4,alfa_range+1):
        dictNACA[num_naca][re]['alfa'].append(float(ws['A'+str(i)].value))
        dictNACA[num_naca][re]['cl'].append(float(ws[chr(96+2*naca).upper() + str(i)].value))
        dictNACA[num_naca][re]['cd'].append(float(ws[chr(96+2*naca+1).upper() + str(i)].value))
logging.info('Fim da importacao dos dados NACA')